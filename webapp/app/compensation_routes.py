"""Compensation module API routes.

Ported from the standalone Solvit Compensation Engine v5 (which was async +
asyncpg) to the portal's synchronous SQLAlchemy session style. All routes are:

  * mounted under the ``/api/comp`` prefix so they never collide with the
    performance portal's own ``/api/solvers`` / ``/api/periods`` endpoints;
  * protected by ``auth.require_admin`` so payroll data inherits the portal
    login (the standalone engine had no auth at all).

The calculation logic itself is untouched — see app/compensation_engine.py.
"""
from __future__ import annotations

import csv
import io
from datetime import date, datetime
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import select, distinct
from sqlalchemy.orm import Session

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app import auth, models
from app.database import get_db
from app.compensation_engine import (
    Params, compute_thresholds, compute_pay, REGION_DATA, MARKET_SHARE,
)

# Every route here requires a valid admin session.
router = APIRouter(
    prefix="/api/comp",
    tags=["compensation"],
    dependencies=[Depends(auth.require_admin)],
)


# ── Pydantic schemas ────────────────────────────────────────────────────────

class ParamsIn(BaseModel):
    rate1: float = 400
    rate2: float = 450
    rate3: float = 500
    stretch_floor: float = 1.15
    wht_pct: float = 5.0
    rev_per_job: float = 950
    t1_day: int = 16
    t2_day: int = 18
    working_days: int = 22
    round_base: int = 5


class SolverUpdate(BaseModel):
    avg_2024: Optional[float] = None
    avg_2025: Optional[float] = None
    avg_2026: Optional[float] = None
    manual_best_override: Optional[float] = None


class PeriodJobIn(BaseModel):
    solver_name: str
    std_jobs: int = 0
    assessment_earnings: float = 0.0


class PeriodComputeIn(BaseModel):
    period_start: date
    period_end: date
    period_label: Optional[str] = None
    jobs: List[PeriodJobIn]
    params: ParamsIn = Field(default_factory=ParamsIn)
    save: bool = False   # persist to DB


# ── Solvers ─────────────────────────────────────────────────────────────────

@router.get("/solvers")
def list_solvers(db: Session = Depends(get_db)):
    """Return all compensation solvers with computed thresholds at default params."""
    solvers = db.execute(
        select(models.SolverCompensation)
        .order_by(models.SolverCompensation.region, models.SolverCompensation.name)
    ).scalars().all()
    params = Params()
    out = []
    for s in solvers:
        tr = compute_thresholds(
            s.avg_2024, s.avg_2025, s.avg_2026,
            s.region, s.manual_best_override, params,
        )
        out.append({
            "id": s.id,
            "name": s.name,
            "region": s.region,
            "avg_2024": s.avg_2024,
            "avg_2025": s.avg_2025,
            "avg_2026": s.avg_2026,
            "active_2026": s.active_2026,
            "manual_best_override": s.manual_best_override,
            **_threshold_dict(tr),
        })
    return out


@router.get("/solvers/{name}/thresholds")
def solver_thresholds(
    name: str,
    params: ParamsIn = Depends(),
    db: Session = Depends(get_db),
):
    """Compute thresholds for a single solver with custom params."""
    s = _get_solver(name, db)
    tr = compute_thresholds(
        s.avg_2024, s.avg_2025, s.avg_2026,
        s.region, s.manual_best_override, Params(**params.model_dump()),
    )
    return {"name": s.name, "region": s.region, **_threshold_dict(tr)}


@router.patch("/solvers/{name}")
def update_solver(name: str, body: SolverUpdate, db: Session = Depends(get_db)):
    """Update historical averages or manual override for a solver."""
    s = _get_solver(name, db)
    changes = body.model_dump(exclude_none=True)
    for k, v in changes.items():
        old = getattr(s, k)
        setattr(s, k, v)
        db.add(models.CompAuditLog(
            action="override_set" if k == "manual_best_override" else "avg_updated",
            solver_name=name,
            old_value=str(old),
            new_value=str(v),
        ))
    s.updated_at = datetime.utcnow()
    db.commit()
    return {"ok": True, "name": name, "updated": list(changes.keys())}


@router.delete("/solvers/{name}/override")
def clear_override(name: str, db: Session = Depends(get_db)):
    """Clear a manual best-average override and revert to the calculated value."""
    s = _get_solver(name, db)
    old = s.manual_best_override
    s.manual_best_override = None
    s.updated_at = datetime.utcnow()
    db.add(models.CompAuditLog(
        action="override_cleared", solver_name=name,
        old_value=str(old), new_value="null",
    ))
    db.commit()
    return {"ok": True, "name": name}


# ── Payroll computation ───────────────────────────────────────────────────────

def _compute_payroll(body: PeriodComputeIn, db: Session) -> dict:
    """Shared computation used by both /compute and /export/csv."""
    params = Params(**body.params.model_dump())
    period_days = (body.period_end - body.period_start).days + 1
    label = body.period_label or _fmt_period(body.period_start, body.period_end)

    solvers = db.execute(select(models.SolverCompensation)).scalars().all()
    solver_map = {s.name: s for s in solvers}

    rows = []
    for job in body.jobs:
        s = solver_map.get(job.solver_name)
        if not s:
            continue
        tr = compute_thresholds(
            s.avg_2024, s.avg_2025, s.avg_2026,
            s.region, s.manual_best_override, params,
        )
        pr = compute_pay(
            job.std_jobs, job.assessment_earnings,
            tr.t1, tr.t2, period_days, params,
        )
        rows.append({
            "name": s.name,
            "region": s.region,
            **_threshold_dict(tr),
            "std_jobs": job.std_jobs,
            "assessment": job.assessment_earnings,
            "gross_pay": pr.gross_pay,
            "wht": pr.wht,
            "net_pay": pr.net_pay,
            "total_gross": pr.total_gross,
            "top_rate": pr.top_rate,
            "band1_jobs": pr.band1_jobs,
            "band2_jobs": pr.band2_jobs,
            "band3_jobs": pr.band3_jobs,
            "t1_period": pr.t1_period,
            "t2_period": pr.t2_period,
        })

        if body.save:
            # Upsert: one row per (solver, period_start).
            existing = db.execute(
                select(models.CompPeriodEntry).where(
                    models.CompPeriodEntry.solver_name == s.name,
                    models.CompPeriodEntry.period_start == body.period_start,
                )
            ).scalar_one_or_none()
            entry = existing or models.CompPeriodEntry(
                solver_name=s.name, period_start=body.period_start,
            )
            entry.period_end = body.period_end
            entry.period_label = label
            entry.std_jobs = job.std_jobs
            entry.assessment_earnings = job.assessment_earnings
            entry.t1_monthly = tr.t1
            entry.t2_monthly = tr.t2
            entry.t1_period = pr.t1_period
            entry.t2_period = pr.t2_period
            entry.gross_pay = pr.gross_pay
            entry.wht = pr.wht
            entry.net_pay = pr.net_pay
            entry.top_rate = pr.top_rate
            if existing is None:
                db.add(entry)

    if body.save:
        db.add(models.CompAuditLog(
            action="period_uploaded", note=f"{label}: {len(rows)} solvers",
        ))
        db.commit()

    total_jobs = sum(r["std_jobs"] for r in rows)
    total_net = sum(r["net_pay"] for r in rows)
    total_gross = sum(r["total_gross"] for r in rows)
    revenue = total_jobs * params.rev_per_job
    margin = ((revenue - total_gross) / revenue * 100) if revenue else 0

    return {
        "period": label,
        "period_start": body.period_start.isoformat(),
        "period_end": body.period_end.isoformat(),
        "period_days": period_days,
        "saved": body.save,
        "summary": {
            "active_solvers": sum(1 for r in rows if r["std_jobs"] > 0),
            "total_jobs": total_jobs,
            "total_gross": round(total_gross, 2),
            "total_net": round(total_net, 2),
            "revenue": round(revenue, 2),
            "direct_margin_pct": round(margin, 2),
        },
        "rows": rows,
    }


@router.post("/compute")
def compute_payroll(body: PeriodComputeIn, db: Session = Depends(get_db)):
    """Compute pay for a list of solver job counts in a given period.
    Optionally persists the results (save=true)."""
    return _compute_payroll(body, db)


# ── Period history ─────────────────────────────────────────────────────────────

@router.get("/periods")
def list_periods(db: Session = Depends(get_db)):
    """List all saved payroll periods."""
    rows = db.execute(
        select(
            models.CompPeriodEntry.period_start,
            models.CompPeriodEntry.period_end,
            models.CompPeriodEntry.period_label,
        ).distinct().order_by(models.CompPeriodEntry.period_start.desc())
    ).all()
    return [
        {"start": r.period_start.isoformat(), "end": r.period_end.isoformat(), "label": r.period_label}
        for r in rows
    ]


@router.get("/periods/{start}/{end}")
def get_period(start: date, end: date, db: Session = Depends(get_db)):
    """Retrieve saved results for a specific period."""
    entries = db.execute(
        select(models.CompPeriodEntry).where(
            models.CompPeriodEntry.period_start == start,
            models.CompPeriodEntry.period_end == end,
        )
    ).scalars().all()
    if not entries:
        raise HTTPException(404, "Period not found")
    return [
        {
            "solver_name": e.solver_name,
            "period_label": e.period_label,
            "std_jobs": e.std_jobs,
            "assessment": float(e.assessment_earnings or 0),
            "gross_pay": float(e.gross_pay or 0),
            "wht": float(e.wht or 0),
            "net_pay": float(e.net_pay or 0),
            "top_rate": e.top_rate,
            "t1_period": e.t1_period,
            "t2_period": e.t2_period,
        }
        for e in entries
    ]


# ── Export ────────────────────────────────────────────────────────────────────

@router.post("/export/csv")
def export_csv(body: PeriodComputeIn, db: Session = Depends(get_db)):
    """Compute and return a payroll CSV download."""
    result = _compute_payroll(body, db)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=[
        "name", "region", "tier", "std_jobs", "t1_period", "t2_period",
        "top_rate", "gross_pay", "assessment", "total_gross", "wht", "net_pay",
    ])
    writer.writeheader()
    for r in result["rows"]:
        writer.writerow({k: r.get(k, "") for k in writer.fieldnames})
    buf.seek(0)
    filename = f"solvit_payroll_{body.period_start}_{body.period_end}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/export/xlsx")
def export_xlsx(body: PeriodComputeIn, db: Session = Depends(get_db)):
    """Server-side Excel export — a styled, multi-sheet .xlsx built with openpyxl.

    Authoritative (computed on the backend, not in the browser). Three sheets:
      1. Billing Schedule — active solvers only, with a totals row, formatted
         for raising invoices.
      2. Full Breakdown   — every solver, all threshold + pay columns.
      3. Summary          — the period headline metrics.
    """
    result = _compute_payroll(body, db)
    label = result["period"]
    days = result["period_days"]
    start, end = result["period_start"], result["period_end"]
    row_map = {r["name"]: r for r in result["rows"]}

    # Pull all solvers so the full-breakdown sheet has the historical averages
    # (the compute rows don't carry avg_2024/25/26).
    solvers = db.execute(
        select(models.SolverCompensation)
        .order_by(models.SolverCompensation.region, models.SolverCompensation.name)
    ).scalars().all()

    # ── Styling helpers ──────────────────────────────────────────────────
    RED = "C0392B"
    BLACK = "1A1A1A"
    title_font = Font(bold=True, size=13, color="FFFFFF")
    hdr_font = Font(bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor=BLACK)
    title_fill = PatternFill("solid", fgColor=RED)
    total_font = Font(bold=True, size=10)
    total_fill = PatternFill("solid", fgColor="F2D7D5")
    money = '#,##0'
    thin = Side(style="thin", color="DDDDDD")
    border = Border(bottom=thin)

    def style_header(ws, row_idx, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = ws.cell(row=row_idx + 1, column=1)

    def autosize(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = openpyxl.Workbook()

    # ── Sheet 1: Billing Schedule ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Billing Schedule"
    ws1.merge_cells("A1:J1")
    ws1["A1"] = "SOLVIT LIMITED — BILLING SCHEDULE"
    ws1["A1"].font = title_font
    ws1["A1"].fill = title_fill
    ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[1].height = 22
    ws1["A2"], ws1["B2"] = "Period:", label
    ws1["A3"], ws1["B3"], ws1["C3"], ws1["D3"], ws1["E3"], ws1["F3"] = \
        "Start:", start, "End:", end, "Days:", days
    ws1["A4"], ws1["B4"] = "Generated:", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    for r in (2, 3, 4):
        ws1.cell(row=r, column=1).font = Font(bold=True, size=10)

    hdr_row = 6
    headers1 = ["Solver Name", "Region", "Tier", "Std Jobs", "Top Rate",
                "Gross Pay", "Assessment", "Total Gross", "WHT (5%)", "Net Pay"]
    for c, h in enumerate(headers1, start=1):
        ws1.cell(row=hdr_row, column=c, value=h)
    style_header(ws1, hdr_row, len(headers1))

    active = [r for r in result["rows"] if r["std_jobs"] > 0 or r["assessment"] > 0]
    active.sort(key=lambda r: r["std_jobs"], reverse=True)
    tot = {"jobs": 0, "gross": 0.0, "assess": 0.0, "total": 0.0, "wht": 0.0, "net": 0.0}
    rr = hdr_row + 1
    for r in active:
        ws1.cell(row=rr, column=1, value=r["name"])
        ws1.cell(row=rr, column=2, value=r["region"])
        ws1.cell(row=rr, column=3, value=r["tier"])
        ws1.cell(row=rr, column=4, value=r["std_jobs"])
        ws1.cell(row=rr, column=5, value=r["top_rate"])
        for col, key, sign in ((6, "gross_pay", 1), (7, "assessment", 1),
                               (8, "total_gross", 1), (9, "wht", -1), (10, "net_pay", 1)):
            cell = ws1.cell(row=rr, column=col, value=round(sign * r[key], 2))
            cell.number_format = money
        for c in range(1, 11):
            ws1.cell(row=rr, column=c).border = border
        tot["jobs"] += r["std_jobs"]; tot["gross"] += r["gross_pay"]
        tot["assess"] += r["assessment"]; tot["total"] += r["total_gross"]
        tot["wht"] += r["wht"]; tot["net"] += r["net_pay"]
        rr += 1

    # Totals row
    ws1.cell(row=rr, column=1, value="TOTAL")
    ws1.cell(row=rr, column=4, value=tot["jobs"])
    for col, val in ((6, tot["gross"]), (7, tot["assess"]), (8, tot["total"]),
                     (9, -tot["wht"]), (10, tot["net"])):
        ws1.cell(row=rr, column=col, value=round(val, 2)).number_format = money
    for c in range(1, 11):
        cell = ws1.cell(row=rr, column=c)
        cell.font = total_font
        cell.fill = total_fill
    autosize(ws1, [22, 14, 8, 9, 9, 13, 12, 13, 12, 13])

    # ── Sheet 2: Full Breakdown ──────────────────────────────────────────
    ws2 = wb.create_sheet("Full Breakdown")
    headers2 = ["Solver", "Region", "Tier", "Avg 2024", "Avg 2025", "Avg 2026",
                "Blended Avg", "Best Average", "Manually Adjusted", "Multiplier",
                "T1", "T2", "T1 (period)", "T2 (period)", "Std Jobs",
                "Band 1", "Band 2", "Band 3", "Top Rate", "Gross Pay",
                "Assessment", "Total Gross", "WHT 5%", "Net Pay", "Basis"]
    for c, h in enumerate(headers2, start=1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(headers2))
    rr = 2
    for s in solvers:
        r = row_map.get(s.name)
        # Threshold values come from the compute row when present; otherwise
        # compute at default params so inactive solvers still show thresholds.
        if r is None:
            tr = compute_thresholds(s.avg_2024, s.avg_2025, s.avg_2026,
                                    s.region, s.manual_best_override, Params(**body.params.model_dump()))
            r = {
                "tier": tr.tier, "blended_avg": tr.blended_avg, "best_average": tr.best_average,
                "is_manually_adjusted": tr.is_manually_adjusted, "multiplier": tr.multiplier,
                "t1": tr.t1, "t2": tr.t2, "t1_period": tr.t1, "t2_period": tr.t2,
                "std_jobs": 0, "band1_jobs": 0, "band2_jobs": 0, "band3_jobs": 0,
                "top_rate": "—", "gross_pay": 0, "assessment": 0, "total_gross": 0,
                "wht": 0, "net_pay": 0, "basis": tr.basis,
            }
        vals = [
            s.name, s.region, r["tier"], s.avg_2024, s.avg_2025, s.avg_2026,
            r["blended_avg"], r["best_average"], "Yes" if r["is_manually_adjusted"] else "No",
            round(r["multiplier"], 2), r["t1"], r["t2"], r["t1_period"], r["t2_period"],
            r["std_jobs"], r["band1_jobs"], r["band2_jobs"], r["band3_jobs"], r["top_rate"],
            r["gross_pay"], r["assessment"], r["total_gross"], r["wht"], r["net_pay"], r["basis"],
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws2.cell(row=rr, column=c, value=v)
            if c in (20, 21, 22, 23, 24):  # money columns
                cell.number_format = money
        rr += 1
    autosize(ws2, [20, 13, 8, 9, 9, 9, 11, 12, 10, 10, 6, 6, 10, 10, 9,
                   8, 8, 8, 9, 12, 11, 12, 11, 12, 40])

    # ── Sheet 3: Summary ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3["A1"] = "Period Summary — " + label
    ws3["A1"].font = Font(bold=True, size=12)
    smry = result["summary"]
    pairs = [
        ("Period", label), ("Start", start), ("End", end), ("Days", days),
        ("Active solvers", smry["active_solvers"]), ("Total jobs", smry["total_jobs"]),
        ("Total gross (KSh)", smry["total_gross"]), ("Total net (KSh)", smry["total_net"]),
        ("Est. revenue (KSh)", smry["revenue"]), ("Direct margin %", smry["direct_margin_pct"]),
    ]
    for i, (k, v) in enumerate(pairs, start=3):
        ws3.cell(row=i, column=1, value=k).font = Font(bold=True, size=10)
        c = ws3.cell(row=i, column=2, value=v)
        if "KSh" in k:
            c.number_format = money
    autosize(ws3, [22, 22])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"solvit_compensation_{start}_to_{end}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Region benchmarks ──────────────────────────────────────────────────────────

@router.get("/benchmarks")
def get_benchmarks():
    """Per-region competitor benchmarks (Regent 2024 × 15% ÷ headcount)."""
    return {
        region: round((d["regent_annual"] / 12 * MARKET_SHARE) / d["solver_count"], 2)
        for region, d in REGION_DATA.items()
    }


# ── Audit log ─────────────────────────────────────────────────────────────────

@router.get("/audit")
def get_audit(limit: int = 50, db: Session = Depends(get_db)):
    rows = db.execute(
        select(models.CompAuditLog)
        .order_by(models.CompAuditLog.created_at.desc())
        .limit(limit)
    ).scalars().all()
    return [
        {
            "id": a.id, "action": a.action, "solver": a.solver_name,
            "old": a.old_value, "new": a.new_value, "note": a.note,
            "at": a.created_at.isoformat() if a.created_at else None,
        }
        for a in rows
    ]


# ── Helpers ─────────────────────────────────────────────────────────────────

def _threshold_dict(tr) -> dict:
    return {
        "blended_avg": tr.blended_avg,
        "region_baseline": tr.region_baseline,
        "best_average": tr.best_average,
        "basis": tr.basis,
        "multiplier": tr.multiplier,
        "active_periods": tr.active_periods,
        "t1": tr.t1,
        "t2": tr.t2,
        "tier": tr.tier,
        "t1_capped": tr.t1_capped,
        "t2_capped": tr.t2_capped,
        "used_region_benchmark": tr.used_region_benchmark,
        "is_manually_adjusted": tr.is_manually_adjusted,
    }


def _get_solver(name: str, db: Session) -> models.SolverCompensation:
    s = db.execute(
        select(models.SolverCompensation).where(models.SolverCompensation.name == name)
    ).scalar_one_or_none()
    if not s:
        raise HTTPException(404, f"Solver not found: {name}")
    return s


def _fmt_period(start: date, end: date) -> str:
    if start.month == end.month and start.year == end.year:
        return start.strftime("%B %Y")
    return f"{start.strftime('%d %b')} - {end.strftime('%d %b %Y')}"
