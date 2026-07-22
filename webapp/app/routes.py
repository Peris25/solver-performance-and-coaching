"""API routes for the portal.

Grouped here rather than spread across files because there aren't enough
endpoints to justify a sub-package. Each route group is clearly labelled.
"""
from __future__ import annotations
from datetime import datetime
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Response
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import select, desc
from pydantic import BaseModel
import io
import re
import shutil

from app import auth, analysis, models, reports
from app.database import get_db
from app.config import settings

router = APIRouter()


# ---------------------------------------------------------------------------
# Health
# ---------------------------------------------------------------------------

@router.get("/api/health")
def health():
    return {"status": "ok", "app": settings.app_name}


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

class LoginPayload(BaseModel):
    username: str
    password: str


@router.post("/api/auth/login")
def login(payload: LoginPayload, response: Response):
    if not auth.authenticate(payload.username, payload.password):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = auth.create_session_token()
    response.set_cookie(
        key=auth.SESSION_COOKIE_NAME,
        value=token,
        max_age=auth.session_max_age_seconds(),
        httponly=True,
        samesite="lax",
        secure=True,  # set True in production behind HTTPS
    )
    return {"ok": True, "user": "admin"}


@router.post("/api/auth/logout")
def logout(response: Response, _: str = Depends(auth.require_admin)):
    response.delete_cookie(auth.SESSION_COOKIE_NAME)
    return {"ok": True}


@router.get("/api/auth/me")
def me(user: str = Depends(auth.require_admin)):
    return {"user": user}


# ---------------------------------------------------------------------------
# Upload — analyse a workbook and persist it as a Period + Snapshots
# ---------------------------------------------------------------------------

@router.post("/api/uploads")
async def upload_workbook(
    file: UploadFile = File(...),
    label: str = Form(...),
    db: Session = Depends(get_db),
    _: str = Depends(auth.require_admin),
):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx/.xls files are accepted")

    # Save the upload to disk (so we can re-analyse later if needed)
    upload_dir = Path(settings.upload_dir)
    upload_dir.mkdir(parents=True, exist_ok=True)
    safe_label = re.sub(r"[^\w.-]+", "_", label).strip("_") or "period"
    timestamp = datetime.utcnow().strftime("%Y%m%dT%H%M%S")
    saved_filename = f"{safe_label}_{timestamp}_{file.filename}"
    saved_path = upload_dir / saved_filename
    with open(saved_path, "wb") as out:
        shutil.copyfileobj(file.file, out)

    # Analyse
    try:
        tv, sb, cr = analysis.read_workbook_sheets(saved_path)
        result = analysis.analyse_dataframes(tv, sb, cr)
    except ValueError as e:
        # Bad workbook shape — surface a clean message
        saved_path.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        saved_path.unlink(missing_ok=True)
        raise HTTPException(status_code=500, detail=f"Failed to analyse: {e}")

    # Upsert Period by label — re-uploading the same period replaces snapshots
    existing = db.scalar(select(models.Period).where(models.Period.label == label))
    if existing:
        # Delete old snapshots; will recreate
        for s in existing.snapshots:
            db.delete(s)
        period = existing
        period.uploaded_at = datetime.utcnow()
        period.uploaded_filename = saved_filename
    else:
        period = models.Period(
            label=label,
            uploaded_at=datetime.utcnow(),
            uploaded_filename=saved_filename,
        )
        db.add(period)
        db.flush()  # get period.id

    # Team fields
    t = result["team"]
    period.total_solvers = t["total_solvers"]
    period.total_valuations = t["total_valuations"]
    period.total_jobs_assigned = t.get("total_jobs_assigned") or 0
    period.total_jobs_pending = t.get("total_jobs_pending") or 0
    period.total_jobs_initiated_by_solvers = t.get("total_jobs_initiated_by_solvers") or 0
    period.team_submission_rate = t["team_submission_rate"]
    period.median_submission_rate = t["median_submission_rate"]
    period.avg_total_tat_hrs = t.get("avg_total_tat_hrs")
    period.median_total_tat_hrs = t.get("median_total_tat_hrs")
    period.avg_response_tat_hrs = t["avg_response_tat_hrs"]
    period.median_response_tat_hrs = t["median_response_tat_hrs"]
    period.avg_onsite_tat_hrs = t["avg_onsite_tat_hrs"]
    period.median_onsite_tat_hrs = t["median_onsite_tat_hrs"]
    period.avg_rating = t["avg_rating"]
    period.avg_volume = t["avg_volume"]
    period.median_volume = t["median_volume"]
    period.pct_stuck_jobs_team = t["pct_stuck_jobs_team"]

    def _parse_iso(v):
        return datetime.fromisoformat(v) if v else None

    period.backlog_period_start = _parse_iso(t.get("backlog_period_start"))
    period.backlog_period_end = _parse_iso(t.get("backlog_period_end"))
    period.total_backlog = t.get("total_backlog") or 0
    period.total_valued_this_period = t.get("total_valued_this_period") or 0
    period.pct_backlog = t.get("pct_backlog")
    period.oldest_backlog_days = t.get("oldest_backlog_days")

    # --- Region matching (must happen before snapshots are saved, so the
    # talent grid can be recomputed with each solver's REGIONAL average
    # jobs-initiated count as the quality-axis benchmark) ---
    registered = db.scalars(select(models.Solver).where(models.Solver.active == 1)).all()
    registered_names = [r.name for r in registered]
    name_to_info = {r.name: {"region": r.region, "location": r.location or ""} for r in registered}

    solver_region_map = {}
    for s in result["solvers"]:
        matched = analysis.find_solver_match(s["name"], registered_names)
        if matched:
            solver_region_map[s["name"]] = name_to_info[matched]
        # else: stays "Unassigned"

    # Recompute talent grids now that regions are known — replaces the
    # provisional team-wide-average version computed inside analyse_workbook.
    analysis.recompute_talent_grids(result["solvers"], solver_region_map)

    # Solver snapshots
    for s in result["solvers"]:
        # Stash the talent-grid result inside `extra` so we don't need a schema migration
        extra_with_grid = dict(s["extra"])
        extra_with_grid["talent_grid"] = s.get("talent_grid")

        snap = models.SolverSnapshot(
            period_id=period.id,
            name=s["name"],
            volume=s["volume"],
            total_attempts=s["total_attempts"],
            valued_count=s["valued_count"],
            assigned_count=s["assigned_count"],
            pending_count=s["pending_count"],
            jobs_initiated=s.get("jobs_initiated", 0),
            stuck_job_count=s["stuck_job_count"],
            n_ratings=s["n_ratings"],
            backlog_count=s.get("backlog_count", 0),
            submission_rate=s["submission_rate"],
            approval_rate=s["approval_rate"],
            avg_total_tat_hrs=s.get("avg_total_tat_hrs"),
            median_total_tat_hrs=s.get("median_total_tat_hrs"),
            avg_response_tat_hrs=s["avg_response_tat_hrs"],
            median_response_tat_hrs=s["median_response_tat_hrs"],
            avg_onsite_tat_hrs=s["avg_onsite_tat_hrs"],
            median_onsite_tat_hrs=s["median_onsite_tat_hrs"],
            avg_rating=s["avg_rating"],
            stuck_job_rate=s["stuck_job_rate"],
            avg_backlog_age_days=s.get("avg_backlog_age_days"),
            classifications=s["classifications"],
            training_modules=s["training_modules"],
            focus_areas=s["focus_areas"],
            extra=extra_with_grid,
        )
        db.add(snap)

    # --- Regional snapshots ---
    # Aggregate by region and persist
    # Delete any old region snapshots for this period before re-inserting
    for r in db.scalars(select(models.RegionSnapshot).where(models.RegionSnapshot.period_id == period.id)).all():
        db.delete(r)

    region_aggs = analysis.aggregate_by_region(result["solvers"], solver_region_map)
    for ra in region_aggs:
        db.add(models.RegionSnapshot(
            period_id=period.id,
            region=ra["region"],
            solver_count=ra["solver_count"],
            total_assigned=ra["total_assigned"],
            total_valued=ra["total_valued"],
            total_pending=ra["total_pending"],
            total_jobs_initiated=ra["total_jobs_initiated"],
            jobs_per_solver=ra["jobs_per_solver"],
            submission_rate=ra["submission_rate"],
            avg_total_tat_hrs=ra["avg_total_tat_hrs"],
            avg_response_tat_hrs=ra["avg_response_tat_hrs"],
            avg_onsite_tat_hrs=ra["avg_onsite_tat_hrs"],
            avg_rating=ra["avg_rating"],
            avg_jobs_initiated=ra["avg_jobs_initiated"],
            staffing=ra["staffing"],
            performance=ra["performance"],
            needs_coaching_count=ra["needs_coaching_count"],
            strong_count=ra["strong_count"],
            locations=ra["locations"],
            solver_names=ra["solver_names"],
        ))

    # --- Row-level persistence for continuous date-range analysis ---
    # (talent grid "1st July to 15th July" style queries). Deduped so
    # re-uploading an overlapping period doesn't double-count a job.
    _persist_job_and_rating_records(db, period.id, tv, sb, cr)

    db.commit()
    db.refresh(period)
    return {
        "ok": True,
        "period_id": period.id,
        "label": period.label,
        "total_solvers": period.total_solvers,
        "total_valuations": period.total_valuations,
    }


def _persist_job_and_rating_records(db: Session, period_id: int, tv, sb, cr) -> None:
    """Bulk-insert raw job/rating rows for date-range queries, skipping any
    whose dedup key already exists (so overlapping uploads — e.g. a weekly
    export whose dates fall inside a later monthly export — don't double-
    count the same job). Not wrapped in its own commit; the caller commits.
    """
    import pandas as pd

    def _clean_dt(v):
        if v is None or (isinstance(v, float) and pd.isna(v)) or pd.isna(v):
            return None
        try:
            return v.to_pydatetime() if hasattr(v, "to_pydatetime") else v
        except (ValueError, TypeError):
            return None

    def _clean_str(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return None
        s = str(v).strip()
        return s or None

    def _clean_float(v):
        try:
            if v is None or pd.isna(v):
                return None
            return float(v)
        except (ValueError, TypeError):
            return None

    # Build candidate JobRecord rows from TOTAL VALUED + SOLVERS BASKET
    job_candidates = []
    for sheet_source, df in (("total_valued", tv), ("solvers_basket", sb)):
        for _, row in df.iterrows():
            solver = _clean_str(row.get("Solver"))
            if not solver:
                continue
            requested_date = _clean_dt(row.get("Requested_Date"))
            vehicle_reg = _clean_str(row.get("Vehicle_reg"))
            dedup_key = analysis.job_record_dedup_key(sheet_source, vehicle_reg, requested_date, solver)
            job_candidates.append({
                "dedup_key": dedup_key,
                "period_id": period_id,
                "sheet_source": sheet_source,
                "solver": solver,
                "vehicle_reg": vehicle_reg,
                "requested_date": requested_date,
                "schedule_date": _clean_dt(row.get("Schedule_date")),
                "valuation_start": _clean_dt(row.get("Valuation_Start")),
                "valuation_date": _clean_dt(row.get("Valuation_Date")),
                "request_status": _clean_str(row.get("Request_Status")),
                "approval_status": _clean_str(row.get("Approval_Status")),
                "initiator_source": _clean_str(row.get("Initiator_Source")),
                "initiated_by": _clean_str(row.get("Initiated_by")),
            })

    if job_candidates:
        keys = [c["dedup_key"] for c in job_candidates]
        existing_keys = set()
        # Chunk the IN clause so very large uploads don't build one giant query
        for i in range(0, len(keys), 1000):
            chunk = keys[i:i + 1000]
            existing_keys.update(
                db.scalars(
                    select(models.JobRecord.dedup_key).where(models.JobRecord.dedup_key.in_(chunk))
                ).all()
            )
        new_jobs = [c for c in job_candidates if c["dedup_key"] not in existing_keys]
        # De-dup within this same batch too (two rows in the same sheet with
        # an identical key would otherwise violate the unique constraint)
        seen = set()
        deduped_new_jobs = []
        for c in new_jobs:
            if c["dedup_key"] in seen:
                continue
            seen.add(c["dedup_key"])
            deduped_new_jobs.append(c)
        if deduped_new_jobs:
            db.bulk_insert_mappings(models.JobRecord, deduped_new_jobs)

    # Build candidate RatingRecord rows from CLIENT RATING
    rating_candidates = []
    for _, row in cr.iterrows():
        solver = _clean_str(row.get("Solver"))
        if not solver:
            continue
        request_id = row.get("Request_ID") if "Request_ID" in cr.columns else None
        dedup_key = analysis.rating_record_dedup_key(request_id)
        rating_candidates.append({
            "dedup_key": dedup_key,
            "period_id": period_id,
            "solver": solver,
            "vehicle_reg": _clean_str(row.get("Vehicle_reg")),
            "initiated_date": _clean_dt(row.get("Initiated_Date")),
            "rating": _clean_float(row.get("rating")),
            "presentation_rating": _clean_float(row.get("presentation_rating")),
            "professionalism_rating": _clean_float(row.get("professionalism_rating")),
            "punctuality_rating": _clean_float(row.get("punctuality_rating")),
        })

    if rating_candidates:
        keys = [c["dedup_key"] for c in rating_candidates]
        existing_keys = set()
        for i in range(0, len(keys), 1000):
            chunk = keys[i:i + 1000]
            existing_keys.update(
                db.scalars(
                    select(models.RatingRecord.dedup_key).where(models.RatingRecord.dedup_key.in_(chunk))
                ).all()
            )
        new_ratings = [c for c in rating_candidates if c["dedup_key"] not in existing_keys]
        seen = set()
        deduped_new_ratings = []
        for c in new_ratings:
            if c["dedup_key"] in seen:
                continue
            seen.add(c["dedup_key"])
            deduped_new_ratings.append(c)
        if deduped_new_ratings:
            db.bulk_insert_mappings(models.RatingRecord, deduped_new_ratings)


# ---------------------------------------------------------------------------
# Periods
# ---------------------------------------------------------------------------

@router.get("/api/periods")
def list_periods(db: Session = Depends(get_db), _: str = Depends(auth.require_admin)):
    rows = db.scalars(select(models.Period).order_by(desc(models.Period.uploaded_at))).all()
    return {"periods": [
        {
            "id": p.id,
            "label": p.label,
            "uploaded_at": p.uploaded_at.isoformat() if p.uploaded_at else None,
            "total_solvers": p.total_solvers,
            "total_valuations": p.total_valuations,
        }
        for p in rows
    ]}


@router.get("/api/periods/{period_id}")
def get_period(period_id: int, db: Session = Depends(get_db), _: str = Depends(auth.require_admin)):
    p = db.get(models.Period, period_id)
    if not p:
        raise HTTPException(status_code=404, detail="Period not found")
    snapshots = sorted(p.snapshots, key=lambda s: s.volume, reverse=True)
    return {
        "id": p.id,
        "label": p.label,
        "uploaded_at": p.uploaded_at.isoformat() if p.uploaded_at else None,
        "team": {
            "total_solvers": p.total_solvers,
            "total_valuations": p.total_valuations,
            "total_jobs_assigned": p.total_jobs_assigned or 0,
            "total_jobs_pending": p.total_jobs_pending or 0,
            "total_jobs_initiated_by_solvers": p.total_jobs_initiated_by_solvers or 0,
            "team_submission_rate": p.team_submission_rate,
            "median_submission_rate": p.median_submission_rate,
            "avg_total_tat_hrs": p.avg_total_tat_hrs,
            "median_total_tat_hrs": p.median_total_tat_hrs,
            "avg_response_tat_hrs": p.avg_response_tat_hrs,
            "median_response_tat_hrs": p.median_response_tat_hrs,
            "avg_onsite_tat_hrs": p.avg_onsite_tat_hrs,
            "median_onsite_tat_hrs": p.median_onsite_tat_hrs,
            "avg_rating": p.avg_rating,
            "avg_volume": p.avg_volume,
            "median_volume": p.median_volume,
            "pct_stuck_jobs_team": p.pct_stuck_jobs_team,
            "backlog_period_start": p.backlog_period_start.isoformat() if p.backlog_period_start else None,
            "backlog_period_end": p.backlog_period_end.isoformat() if p.backlog_period_end else None,
            "total_backlog": p.total_backlog or 0,
            "total_valued_this_period": p.total_valued_this_period or 0,
            "pct_backlog": p.pct_backlog,
            "oldest_backlog_days": p.oldest_backlog_days,
        },
        "targets": analysis.TARGETS,
        "solvers": [snapshot_to_dict(s) for s in snapshots],
    }


@router.delete("/api/periods/{period_id}")
def delete_period(period_id: int, db: Session = Depends(get_db), _: str = Depends(auth.require_admin)):
    """Delete an uploaded period entirely: its snapshots, region snapshots,
    email send log, row-level job/rating records, and the saved .xlsx file
    on disk. This can't be undone — the admin has to re-upload the workbook
    to get this period back.
    """
    p = db.get(models.Period, period_id)
    if not p:
        raise HTTPException(status_code=404, detail="Period not found")

    label = p.label
    uploaded_filename = p.uploaded_filename

    # SolverSnapshot rows cascade automatically via the relationship, but
    # RegionSnapshot, EmailSendLog, JobRecord, and RatingRecord don't have a
    # cascading relationship defined, so clean those up explicitly first.
    for r in db.scalars(select(models.RegionSnapshot).where(models.RegionSnapshot.period_id == period_id)).all():
        db.delete(r)
    for log in db.scalars(select(models.EmailSendLog).where(models.EmailSendLog.period_id == period_id)).all():
        db.delete(log)
    for jr in db.scalars(select(models.JobRecord).where(models.JobRecord.period_id == period_id)).all():
        db.delete(jr)
    for rr in db.scalars(select(models.RatingRecord).where(models.RatingRecord.period_id == period_id)).all():
        db.delete(rr)

    db.delete(p)
    db.commit()

    # Best-effort removal of the underlying uploaded workbook from disk.
    # Not fatal if this fails (e.g. already gone) — the DB rows are the
    # source of truth for what the portal shows.
    file_removed = False
    if uploaded_filename:
        file_path = Path(settings.upload_dir) / uploaded_filename
        try:
            if file_path.exists():
                file_path.unlink()
                file_removed = True
        except OSError:
            pass

    return {"ok": True, "deleted_label": label, "file_removed": file_removed}


# ---------------------------------------------------------------------------
# Solvers — per-period & history
# ---------------------------------------------------------------------------

def snapshot_to_dict(s: models.SolverSnapshot) -> dict:
    extra = s.extra or {}
    return {
        "name": s.name,
        "volume": s.volume,
        "total_attempts": s.total_attempts,
        "valued_count": s.valued_count,
        "assigned_count": s.assigned_count,
        "pending_count": s.pending_count,
        "jobs_initiated": s.jobs_initiated or 0,
        "stuck_job_count": s.stuck_job_count,
        "n_ratings": s.n_ratings,
        "backlog_count": s.backlog_count or 0,
        "avg_backlog_age_days": s.avg_backlog_age_days,
        "submission_rate": s.submission_rate,
        "approval_rate": s.approval_rate,
        "avg_total_tat_hrs": s.avg_total_tat_hrs,
        "median_total_tat_hrs": s.median_total_tat_hrs,
        "avg_response_tat_hrs": s.avg_response_tat_hrs,
        "median_response_tat_hrs": s.median_response_tat_hrs,
        "avg_onsite_tat_hrs": s.avg_onsite_tat_hrs,
        "median_onsite_tat_hrs": s.median_onsite_tat_hrs,
        "avg_rating": s.avg_rating,
        "stuck_job_rate": s.stuck_job_rate,
        "classifications": s.classifications or {},
        "training_modules": s.training_modules or [],
        "focus_areas": s.focus_areas or [],
        "talent_grid": extra.get("talent_grid"),
        "extra": extra,
    }


@router.get("/api/solvers/{name}/history")
def solver_history(name: str, db: Session = Depends(get_db), _: str = Depends(auth.require_admin)):
    """All snapshots for a solver across periods, oldest first.

    Used for month-over-month trend charts in the UI.
    """
    snapshots = db.scalars(
        select(models.SolverSnapshot)
        .where(models.SolverSnapshot.name == name)
        .join(models.Period)
        .order_by(models.Period.uploaded_at.asc())
    ).all()
    return {
        "name": name,
        "periods": [
            {
                "period_id": s.period_id,
                "period_label": s.period.label,
                "uploaded_at": s.period.uploaded_at.isoformat() if s.period.uploaded_at else None,
                **snapshot_to_dict(s),
            }
            for s in snapshots
        ],
    }


# ---------------------------------------------------------------------------
# Reports — generate a per-solver Word doc on demand
# ---------------------------------------------------------------------------

class IntroPayload(BaseModel):
    intro: Optional[str] = None


@router.post("/api/periods/{period_id}/solvers/{name}/report")
def generate_report(
    period_id: int,
    name: str,
    payload: IntroPayload = IntroPayload(),
    compare_to: Optional[int] = None,
    db: Session = Depends(get_db),
    _: str = Depends(auth.require_admin),
):
    period = db.get(models.Period, period_id)
    if not period:
        raise HTTPException(status_code=404, detail="Period not found")
    snap = db.scalar(
        select(models.SolverSnapshot)
        .where(
            models.SolverSnapshot.period_id == period_id,
            models.SolverSnapshot.name == name,
        )
    )
    if not snap:
        raise HTTPException(status_code=404, detail="Solver not in this period")

    # Optional: pull comparison snapshot from a previous period
    comparison_data = None
    previous_period_label = None
    if compare_to:
        prev_period = db.get(models.Period, compare_to)
        prev_snap = db.scalar(
            select(models.SolverSnapshot).where(
                models.SolverSnapshot.period_id == compare_to,
                models.SolverSnapshot.name == name,
            )
        )
        if prev_period and prev_snap:
            comparison_data = analysis.compare_snapshots(
                snapshot_to_dict(snap),
                snapshot_to_dict(prev_snap),
            )
            previous_period_label = prev_period.label

    docx_bytes = reports.build_doc(
        snapshot_to_dict(snap),
        team={
            "total_solvers": period.total_solvers,
            "total_valuations": period.total_valuations,
            "team_submission_rate": period.team_submission_rate,
            "median_submission_rate": period.median_submission_rate,
            "avg_total_tat_hrs": period.avg_total_tat_hrs,
            "median_total_tat_hrs": period.median_total_tat_hrs,
            "avg_response_tat_hrs": period.avg_response_tat_hrs,
            "median_response_tat_hrs": period.median_response_tat_hrs,
            "avg_onsite_tat_hrs": period.avg_onsite_tat_hrs,
            "median_onsite_tat_hrs": period.median_onsite_tat_hrs,
            "avg_rating": period.avg_rating,
            "avg_volume": period.avg_volume,
            "median_volume": period.median_volume,
            "pct_stuck_jobs_team": period.pct_stuck_jobs_team,
        },
        targets=analysis.TARGETS,
        period_label=period.label,
        personalised_intro=payload.intro,
        comparison=comparison_data,
        previous_period_label=previous_period_label,
    )

    safe_name = re.sub(r"[^\w]+", "_", name).strip("_")
    filename = f"{safe_name}_coaching_{period.label.replace(' ', '_')}.docx"

    return StreamingResponse(
        io.BytesIO(docx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Solver emails — admin manages the name → email mapping here
# ---------------------------------------------------------------------------

class SolverEmailPayload(BaseModel):
    email: str


@router.get("/api/solver-emails")
def list_solver_emails(db: Session = Depends(get_db), _: str = Depends(auth.require_admin)):
    rows = db.scalars(select(models.SolverEmail).order_by(models.SolverEmail.name)).all()
    return {"emails": [{"name": r.name, "email": r.email,
                        "updated_at": r.updated_at.isoformat()} for r in rows]}


@router.put("/api/solver-emails/{name}")
def upsert_solver_email(
    name: str,
    payload: SolverEmailPayload,
    db: Session = Depends(get_db),
    _: str = Depends(auth.require_admin),
):
    if "@" not in payload.email or "." not in payload.email.split("@")[-1]:
        raise HTTPException(status_code=400, detail="Invalid email address")

    existing = db.scalar(select(models.SolverEmail).where(models.SolverEmail.name == name))
    if existing:
        existing.email = payload.email
    else:
        db.add(models.SolverEmail(name=name, email=payload.email))
    db.commit()
    return {"ok": True, "name": name, "email": payload.email}


@router.delete("/api/solver-emails/{name}")
def delete_solver_email(name: str, db: Session = Depends(get_db),
                        _: str = Depends(auth.require_admin)):
    row = db.scalar(select(models.SolverEmail).where(models.SolverEmail.name == name))
    if row:
        db.delete(row)
        db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Email coaching reports — single send + bulk send to "didn't meet target"
# ---------------------------------------------------------------------------

class EmailSendResult(BaseModel):
    name: str
    status: str  # "sent" | "skipped_no_email" | "failed" | "skipped_strong"
    detail: Optional[str] = None


def _build_doc_for(period: models.Period, snap: models.SolverSnapshot) -> bytes:
    """Build the .docx bytes for one solver from their snapshot."""
    return reports.build_doc(
        snapshot_to_dict(snap),
        team={
            "total_solvers": period.total_solvers,
            "total_valuations": period.total_valuations,
            "team_submission_rate": period.team_submission_rate,
            "median_submission_rate": period.median_submission_rate,
            "avg_total_tat_hrs": period.avg_total_tat_hrs,
            "median_total_tat_hrs": period.median_total_tat_hrs,
            "avg_response_tat_hrs": period.avg_response_tat_hrs,
            "median_response_tat_hrs": period.median_response_tat_hrs,
            "avg_onsite_tat_hrs": period.avg_onsite_tat_hrs,
            "median_onsite_tat_hrs": period.median_onsite_tat_hrs,
            "avg_rating": period.avg_rating,
            "avg_volume": period.avg_volume,
            "median_volume": period.median_volume,
            "pct_stuck_jobs_team": period.pct_stuck_jobs_team,
        },
        targets=analysis.TARGETS,
        period_label=period.label,
        personalised_intro=None,
    )


def _email_one(period: models.Period, snap: models.SolverSnapshot,
               to_email: str, trigger: str, db: Session,
               custom_body: str | None = None) -> EmailSendResult:
    """Send the coaching email to a single solver and record the result."""
    import app.emails as emails  # local import so the module can be missing in dev

    log = models.EmailSendLog(
        period_id=period.id,
        solver_name=snap.name,
        email_to=to_email,
        status="sent",
        trigger=trigger,
        focus_areas=snap.focus_areas or [],
    )
    try:
        docx_bytes = _build_doc_for(period, snap)
        emails.send_coaching_email(
            to_email=to_email,
            solver_name=snap.name,
            period_label=period.label,
            stats=snapshot_to_dict(snap),
            docx_bytes=docx_bytes,
            targets=analysis.TARGETS,
            custom_body=custom_body,
        )
        log.status = "sent"
        db.add(log)
        db.commit()
        return EmailSendResult(name=snap.name, status="sent", detail=to_email)
    except emails.EmailNotConfigured as e:
        log.status = "failed"
        log.error = str(e)
        db.add(log)
        db.commit()
        return EmailSendResult(name=snap.name, status="failed", detail=str(e))
    except Exception as e:
        log.status = "failed"
        log.error = f"{type(e).__name__}: {e}"
        db.add(log)
        db.commit()
        return EmailSendResult(name=snap.name, status="failed", detail=str(e))


class EmailPayload(BaseModel):
    custom_body: Optional[str] = None  # admin-edited text; None = auto-generate


@router.get("/api/periods/{period_id}/solvers/{name}/email-preview")
def email_preview(
    period_id: int,
    name: str,
    db: Session = Depends(get_db),
    _: str = Depends(auth.require_admin),
):
    """Return the subject + body that would be sent, ready for the admin to edit."""
    import app.emails as emails
    period = db.get(models.Period, period_id)
    if not period:
        raise HTTPException(status_code=404, detail="Period not found")
    snap = db.scalar(select(models.SolverSnapshot).where(
        models.SolverSnapshot.period_id == period_id,
        models.SolverSnapshot.name == name,
    ))
    if not snap:
        raise HTTPException(status_code=404, detail="Solver not in this period")
    email_row = db.scalar(select(models.SolverEmail).where(models.SolverEmail.name == name))
    preview = emails.build_preview(snapshot_to_dict(snap), period.label, analysis.TARGETS)
    preview["email_on_file"] = email_row.email if email_row else None
    preview["name"] = name
    return preview


@router.get("/api/periods/{period_id}/email-template/{focus}")
def email_template(
    period_id: int,
    focus: str,
    db: Session = Depends(get_db),
    _: str = Depends(auth.require_admin),
):
    """Return an editable bulk template for a focus area (time / submission / recognition).

    The template uses {{tokens}} that are replaced per-solver at send time.
    The admin edits this once and it applies to every solver with that focus.
    """
    import app.emails as emails
    if focus not in ("time", "submission", "recognition"):
        raise HTTPException(status_code=400, detail="focus must be time, submission, or recognition")
    period = db.get(models.Period, period_id)
    if not period:
        raise HTTPException(status_code=404, detail="Period not found")

    tpl = emails.build_template(focus, analysis.TARGETS)

    # Count how many solvers would receive this
    snaps = period.snapshots
    if focus == "recognition":
        count = sum(1 for s in snaps if (s.focus_areas or []) == ["strong"] or "strong" in (s.focus_areas or []))
    else:
        count = sum(1 for s in snaps if focus in (s.focus_areas or []))
    tpl["solver_count"] = count
    tpl["period_label"] = period.label
    tpl["token_help"] = (
        "{{name}} full name · {{first_name}} first name only · {{volume}} jobs completed · "
        "{{avg_tat}} avg TAT · {{tat_target}} target · {{assigned}} assigned · "
        "{{completed}} completed · {{pending}} pending · {{submission_pct}} sub% · "
        "{{rating}} client rating · {{period_label}} period name"
    )
    return tpl


@router.post("/api/periods/{period_id}/solvers/{name}/email")

def email_one_solver(
    period_id: int,
    name: str,
    payload: EmailPayload = EmailPayload(),
    db: Session = Depends(get_db),
    _: str = Depends(auth.require_admin),
):
    """Send the coaching/recognition email for one solver.

    Pass `custom_body` in the JSON body to override the auto-generated text.
    """
    period = db.get(models.Period, period_id)
    if not period:
        raise HTTPException(status_code=404, detail="Period not found")
    snap = db.scalar(select(models.SolverSnapshot).where(
        models.SolverSnapshot.period_id == period_id,
        models.SolverSnapshot.name == name,
    ))
    if not snap:
        raise HTTPException(status_code=404, detail="Solver not in this period")
    email_row = db.scalar(select(models.SolverEmail).where(models.SolverEmail.name == name))
    if not email_row:
        raise HTTPException(
            status_code=400,
            detail=f"No email on file for {name}. Add one in Solver Emails first.",
        )

    result = _email_one(
        period, snap, email_row.email,
        trigger="manual",
        db=db,
        custom_body=payload.custom_body,
    )
    if result.status == "failed":
        raise HTTPException(status_code=500, detail=result.detail)
    return result

class BulkEmailPayload(BaseModel):
    # Optional edited templates keyed by focus area: {"time": "...", "submission": "...", "recognition": "..."}
    # Values are plain-text bodies with {{tokens}} already resolved or left for per-solver resolution.
    custom_templates: dict[str, str] = {}

@router.post("/api/periods/{period_id}/send-coaching-emails")
def bulk_send_coaching_emails(
    period_id: int,
    payload: BulkEmailPayload = BulkEmailPayload(),
    db: Session = Depends(get_db),
    _: str = Depends(auth.require_admin),
):
    """Send coaching emails to every solver who didn't meet their targets this period.

    Pass `custom_templates` to override the auto-generated body per focus area.
    Templates may contain {{tokens}} — they are resolved per-solver before sending.
    """
    import app.emails as emails
    period = db.get(models.Period, period_id)
    if not period:
        raise HTTPException(status_code=404, detail="Period not found")

    snaps = sorted(period.snapshots, key=lambda s: s.name)
    email_map = {
        r.name: r.email for r in
        db.scalars(select(models.SolverEmail)).all()
    }

    results: list[EmailSendResult] = []
    for snap in snaps:
        focus = snap.focus_areas or []
        if not focus or focus == ["strong"]:
            results.append(EmailSendResult(
                name=snap.name, status="skipped_strong",
                detail="no focus areas flagged"
            ))
            continue
        to_email = email_map.get(snap.name)
        if not to_email:
            results.append(EmailSendResult(
                name=snap.name, status="skipped_no_email",
                detail="add email in Solver Emails"
            ))
            continue
        # Resolve custom template if provided: pick the most specific focus area
        custom_body = None
        for f in focus:
            if f in payload.custom_templates:
                tpl = payload.custom_templates[f]
                custom_body = emails.apply_template(tpl, snapshot_to_dict(snap), period.label)
                break
        result = _email_one(
            period, snap, to_email,
            trigger="bulk_targets_missed",
            db=db,
            custom_body=custom_body,
        )
        results.append(result)

    return {
        "ok": True,
        "period_id": period_id,
        "results": [r.dict() for r in results],
        "summary": {
            "sent": sum(1 for r in results if r.status == "sent"),
            "failed": sum(1 for r in results if r.status == "failed"),
            "skipped_no_email": sum(1 for r in results if r.status == "skipped_no_email"),
            "skipped_strong": sum(1 for r in results if r.status == "skipped_strong"),
        },
    }


@router.post("/api/periods/{period_id}/send-recognition-emails")
def bulk_send_recognition_emails(
    period_id: int,
    payload: BulkEmailPayload = BulkEmailPayload(),
    db: Session = Depends(get_db),
    _: str = Depends(auth.require_admin),
):
    """Send recognition emails to every solver who hit all targets this period.

    Pass `custom_templates` with key "recognition" to override the body.
    """
    import app.emails as emails
    period = db.get(models.Period, period_id)
    if not period:
        raise HTTPException(status_code=404, detail="Period not found")

    snaps = sorted(period.snapshots, key=lambda s: s.name)
    email_map = {
        r.name: r.email for r in
        db.scalars(select(models.SolverEmail)).all()
    }

    results: list[EmailSendResult] = []
    for snap in snaps:
        focus = snap.focus_areas or []
        # Only strong performers — skip anyone with a needs_work classification
        if "strong" not in focus:
            results.append(EmailSendResult(
                name=snap.name, status="skipped_not_strong",
                detail="not a strong performer this period"
            ))
            continue
        to_email = email_map.get(snap.name)
        if not to_email:
            results.append(EmailSendResult(
                name=snap.name, status="skipped_no_email",
                detail="add email in Solver Emails"
            ))
            continue
        result = _email_one(
            period, snap, to_email,
            trigger="bulk_recognition",
            db=db,
            custom_body=emails.apply_template(
                payload.custom_templates["recognition"],
                snapshot_to_dict(snap), period.label
            ) if "recognition" in payload.custom_templates else None,
        )
        results.append(result)

    return {
        "ok": True,
        "period_id": period_id,
        "results": [r.dict() for r in results],
        "summary": {
            "sent": sum(1 for r in results if r.status == "sent"),
            "failed": sum(1 for r in results if r.status == "failed"),
            "skipped_no_email": sum(1 for r in results if r.status == "skipped_no_email"),
            "skipped_not_strong": sum(1 for r in results if r.status == "skipped_not_strong"),
        },
    }


@router.get("/api/periods/{period_id}/email-log")
def period_email_log(
    period_id: int,
    db: Session = Depends(get_db),
    _: str = Depends(auth.require_admin),
):
    rows = db.scalars(
        select(models.EmailSendLog)
        .where(models.EmailSendLog.period_id == period_id)
        .order_by(desc(models.EmailSendLog.sent_at))
    ).all()
    return {
        "log": [{
            "id": r.id,
            "solver_name": r.solver_name,
            "email_to": r.email_to,
            "sent_at": r.sent_at.isoformat(),
            "status": r.status,
            "trigger": r.trigger,
            "error": r.error,
            "focus_areas": r.focus_areas,
        } for r in rows]
    }


# ---------------------------------------------------------------------------
# Registered solvers (CRUD) — the canonical roster with region/location.
# Used for regional analysis. Loaded once from SOLVERS_REGIONAL_LIST.xlsx
# and then maintained via this UI as new hires come on.
# ---------------------------------------------------------------------------

class SolverPayload(BaseModel):
    name: str
    region: str
    location: Optional[str] = None
    is_lead: Optional[bool] = False


def _solver_to_dict(s: models.Solver) -> dict:
    return {
        "id": s.id,
        "name": s.name,
        "region": s.region,
        "location": s.location or "",
        "active": bool(s.active),
        "is_lead": bool(s.is_lead),
        "created_at": s.created_at.isoformat() if s.created_at else None,
    }


@router.get("/api/registered-solvers")
def list_registered_solvers(db: Session = Depends(get_db), _: str = Depends(auth.require_admin)):
    rows = db.scalars(select(models.Solver).order_by(models.Solver.region, models.Solver.name)).all()
    return {"solvers": [_solver_to_dict(s) for s in rows]}


@router.post("/api/registered-solvers")
def create_registered_solver(
    payload: SolverPayload,
    db: Session = Depends(get_db),
    _: str = Depends(auth.require_admin),
):
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")
    region = payload.region.strip()
    if not region:
        raise HTTPException(status_code=400, detail="Region is required")

    existing = db.scalar(select(models.Solver).where(models.Solver.name == name))
    if existing:
        raise HTTPException(status_code=400, detail=f"Solver {name!r} already exists")

    s = models.Solver(
        name=name,
        region=region,
        location=(payload.location or "").strip(),
        is_lead=1 if payload.is_lead else 0,
    )
    db.add(s)
    db.commit()
    db.refresh(s)
    return _solver_to_dict(s)


@router.put("/api/registered-solvers/{solver_id}")
def update_registered_solver(
    solver_id: int,
    payload: SolverPayload,
    db: Session = Depends(get_db),
    _: str = Depends(auth.require_admin),
):
    s = db.get(models.Solver, solver_id)
    if not s:
        raise HTTPException(status_code=404, detail="Solver not found")
    s.name = payload.name.strip() or s.name
    s.region = payload.region.strip() or s.region
    s.location = (payload.location or "").strip()
    s.is_lead = 1 if payload.is_lead else 0
    db.commit()
    db.refresh(s)
    return _solver_to_dict(s)


@router.delete("/api/registered-solvers/{solver_id}")
def delete_registered_solver(
    solver_id: int,
    db: Session = Depends(get_db),
    _: str = Depends(auth.require_admin),
):
    s = db.get(models.Solver, solver_id)
    if not s:
        raise HTTPException(status_code=404, detail="Solver not found")
    db.delete(s)
    db.commit()
    return {"ok": True}


@router.post("/api/registered-solvers/seed")
async def seed_registered_solvers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: str = Depends(auth.require_admin),
):
    """Seed the solvers table from a SOLVERS_REGIONAL_LIST.xlsx file.

    Parses the regional list (rows have name in col A, location in col B,
    with region header rows where col B == 'Location'). Stars on names
    indicate regional leads.

    Upserts: existing solvers (by name) get region/location updated,
    new ones are inserted.
    """
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Upload an Excel file")

    from openpyxl import load_workbook
    import tempfile
    tmp = Path(tempfile.gettempdir()) / f"seed_{datetime.utcnow().timestamp()}.xlsx"
    with open(tmp, "wb") as out:
        shutil.copyfileobj(file.file, out)

    try:
        wb = load_workbook(tmp, read_only=True)
    except Exception as e:
        tmp.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail=f"Could not read Excel: {e}")

    ws = wb.active
    current_region = None
    parsed: list[tuple[str, str, str, bool]] = []  # (name, region, location, is_lead)

    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        a = str(row[0]).strip()
        b = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

        # Region header row: col B says "Location"
        if b.lower() == "location":
            current_region = a
            continue
        if not current_region:
            continue
        # Skip header text rows
        if a.lower().startswith("detailed solver") or a.lower() == "name":
            continue

        is_lead = a.startswith("⭐")
        name = a.lstrip("⭐ ").strip()
        if not name:
            continue
        parsed.append((name, current_region, b, is_lead))

    tmp.unlink(missing_ok=True)

    added, updated, skipped_duplicates = 0, 0, 0
    seen_in_batch: set[str] = set()
    for name, region, location, is_lead in parsed:
        # Skip if we've already processed this name in this seed pass
        # (handles solvers listed under multiple regions in the source spreadsheet —
        # we keep the FIRST occurrence; admin can edit afterwards)
        if name in seen_in_batch:
            skipped_duplicates += 1
            continue
        seen_in_batch.add(name)

        existing = db.scalar(select(models.Solver).where(models.Solver.name == name))
        if existing:
            existing.region = region
            existing.location = location
            existing.is_lead = 1 if is_lead else 0
            updated += 1
        else:
            db.add(models.Solver(
                name=name, region=region, location=location,
                is_lead=1 if is_lead else 0, active=1,
            ))
            added += 1

    db.commit()
    return {
        "ok": True,
        "added": added,
        "updated": updated,
        "skipped_duplicates": skipped_duplicates,
        "total_parsed": len(parsed),
    }


# ---------------------------------------------------------------------------
# Talent grid — continuous date-range analysis (not tied to any one upload).
# ---------------------------------------------------------------------------

@router.get("/api/talent-grid/range")
def talent_grid_range(
    start: str,
    end: str,
    db: Session = Depends(get_db),
    _: str = Depends(auth.require_admin),
):
    """9-box talent grid computed over an arbitrary date window, e.g.
    start=2026-07-01&end=2026-07-15 — spans across whichever uploads happen
    to contain jobs in that window, rather than being locked to one period.

    Job rows are matched to the window by `valuation_date` (when the job
    was actually completed) — a job assigned June 28th but finished July 2nd
    counts toward a "July 1–15" query, since that's when the solver's work
    on it actually landed. Basket rows that are still pending (no
    valuation_date yet) are matched on `requested_date` instead, so an
    admin looking at "jobs assigned this window" still sees them for the
    submission-rate denominator.

    Returns the same shape as a period's solver list — including
    `talent_grid` per solver — so the frontend can reuse its existing
    rendering code for either a period or a range.
    """
    try:
        start_dt = datetime.strptime(start, "%Y-%m-%d")
        end_dt = datetime.strptime(end, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="start/end must be YYYY-MM-DD")
    if end_dt < start_dt:
        raise HTTPException(status_code=400, detail="end must be on or after start")
    # Make the end date inclusive of the whole day
    end_dt = end_dt.replace(hour=23, minute=59, second=59)

    from sqlalchemy import or_, and_

    job_rows = db.scalars(
        select(models.JobRecord).where(
            or_(
                and_(models.JobRecord.valuation_date.isnot(None),
                     models.JobRecord.valuation_date >= start_dt,
                     models.JobRecord.valuation_date <= end_dt),
                and_(models.JobRecord.valuation_date.is_(None),
                     models.JobRecord.requested_date.isnot(None),
                     models.JobRecord.requested_date >= start_dt,
                     models.JobRecord.requested_date <= end_dt),
            )
        )
    ).all()

    rating_rows = db.scalars(
        select(models.RatingRecord).where(
            models.RatingRecord.initiated_date.isnot(None),
            models.RatingRecord.initiated_date >= start_dt,
            models.RatingRecord.initiated_date <= end_dt,
        )
    ).all()

    if not job_rows:
        return {
            "start": start, "end": end,
            "solver_count": 0, "solvers": [], "team": None,
            "targets": analysis.TARGETS,
            "note": "No job records fall in this date range. Row-level data "
                    "is only captured for workbooks uploaded after this "
                    "feature was added — older periods need to be re-uploaded "
                    "once to backfill it.",
        }

    job_dicts = [{
        "sheet_source": j.sheet_source, "solver": j.solver, "vehicle_reg": j.vehicle_reg,
        "requested_date": j.requested_date, "schedule_date": j.schedule_date,
        "valuation_start": j.valuation_start, "valuation_date": j.valuation_date,
        "request_status": j.request_status, "approval_status": j.approval_status,
        "initiator_source": j.initiator_source, "initiated_by": j.initiated_by,
    } for j in job_rows]
    rating_dicts = [{
        "solver": r.solver, "vehicle_reg": r.vehicle_reg, "initiated_date": r.initiated_date,
        "rating": r.rating, "presentation_rating": r.presentation_rating,
        "professionalism_rating": r.professionalism_rating, "punctuality_rating": r.punctuality_rating,
    } for r in rating_rows]

    tv, sb, cr = analysis.records_to_dataframes(job_dicts, rating_dicts)
    result = analysis.analyse_dataframes(tv, sb, cr)

    # Region-aware talent grid, same as the upload path
    registered = db.scalars(select(models.Solver).where(models.Solver.active == 1)).all()
    registered_names = [r.name for r in registered]
    name_to_info = {r.name: {"region": r.region, "location": r.location or ""} for r in registered}
    solver_region_map = {}
    for s in result["solvers"]:
        matched = analysis.find_solver_match(s["name"], registered_names)
        if matched:
            solver_region_map[s["name"]] = name_to_info[matched]
    analysis.recompute_talent_grids(result["solvers"], solver_region_map)

    return {
        "start": start, "end": end,
        "solver_count": len(result["solvers"]),
        "solvers": result["solvers"],
        "team": result["team"],
        "targets": analysis.TARGETS,
    }


# ---------------------------------------------------------------------------
# Regions — per-period regional aggregations (for the map and tiles).
# ---------------------------------------------------------------------------

@router.get("/api/periods/{period_id}/regions")
def get_period_regions(
    period_id: int,
    recompute: bool = False,
    db: Session = Depends(get_db),
    _: str = Depends(auth.require_admin),
):
    """Return all region snapshots for a period. The dashboard uses these
    for the regional tiles and the Kenya map.

    `recompute=true` forces re-aggregation from the live solver list (useful
    after seeding solvers or adding/removing registered solvers).
    """
    p = db.get(models.Period, period_id)
    if not p:
        raise HTTPException(status_code=404, detail="Period not found")

    rows = db.scalars(
        select(models.RegionSnapshot)
        .where(models.RegionSnapshot.period_id == period_id)
        .order_by(desc(models.RegionSnapshot.total_assigned))
    ).all()

    # Detect "stale" stored data: if Unassigned dominates, the registered
    # solver list has likely changed since upload — recompute.
    stale = False
    if rows:
        unassigned_row = next((r for r in rows if r.region == "Unassigned"), None)
        total_solvers = sum(r.solver_count for r in rows)
        if unassigned_row and total_solvers > 0:
            unassigned_pct = unassigned_row.solver_count / total_solvers
            if unassigned_pct > 0.5:
                stale = True

    # Recompute path: either explicitly requested, stored data is stale,
    # or there are no stored region snapshots at all
    if recompute or stale or not rows:
        registered = db.scalars(select(models.Solver).where(models.Solver.active == 1)).all()
        name_to_info = {r.name: {"region": r.region, "location": r.location or ""} for r in registered}
        registered_names = [r.name for r in registered]

        solver_region_map = {}
        snap_dicts = [snapshot_to_dict(s) for s in p.snapshots]
        for s in snap_dicts:
            matched = analysis.find_solver_match(s["name"], registered_names)
            if matched:
                solver_region_map[s["name"]] = name_to_info[matched]

        regions = analysis.aggregate_by_region(snap_dicts, solver_region_map)

        # Regions may have changed since upload (solver moved / roster edited) —
        # recompute each solver's talent grid with the fresh regional benchmark
        # and re-persist it into their snapshot's `extra` JSON.
        analysis.recompute_talent_grids(snap_dicts, solver_region_map)
        snap_by_name = {s.name: s for s in p.snapshots}
        for sd in snap_dicts:
            snap_row = snap_by_name.get(sd["name"])
            if snap_row is not None:
                extra = dict(snap_row.extra or {})
                extra["talent_grid"] = sd.get("talent_grid")
                snap_row.extra = extra

        # Persist the recomputed region snapshots (overwrite existing)
        for r in rows:
            db.delete(r)
        for ra in regions:
            db.add(models.RegionSnapshot(
                period_id=p.id,
                region=ra["region"],
                solver_count=ra["solver_count"],
                total_assigned=ra["total_assigned"],
                total_valued=ra["total_valued"],
                total_pending=ra["total_pending"],
                total_jobs_initiated=ra["total_jobs_initiated"],
                jobs_per_solver=ra["jobs_per_solver"],
                submission_rate=ra["submission_rate"],
                avg_total_tat_hrs=ra["avg_total_tat_hrs"],
                avg_response_tat_hrs=ra["avg_response_tat_hrs"],
                avg_onsite_tat_hrs=ra["avg_onsite_tat_hrs"],
                avg_rating=ra["avg_rating"],
                avg_jobs_initiated=ra["avg_jobs_initiated"],
                staffing=ra["staffing"],
                performance=ra["performance"],
                needs_coaching_count=ra["needs_coaching_count"],
                strong_count=ra["strong_count"],
                locations=ra["locations"],
                solver_names=ra["solver_names"],
            ))
        db.commit()
        return {"period_id": period_id, "regions": regions}

    return {
        "period_id": period_id,
        "regions": [
            {
                "region": r.region,
                "solver_count": r.solver_count,
                "total_assigned": r.total_assigned,
                "total_valued": r.total_valued,
                "total_pending": r.total_pending,
                "total_jobs_initiated": r.total_jobs_initiated,
                "jobs_per_solver": r.jobs_per_solver,
                "submission_rate": r.submission_rate,
                "avg_total_tat_hrs": r.avg_total_tat_hrs,
                "avg_response_tat_hrs": r.avg_response_tat_hrs,
                "avg_onsite_tat_hrs": r.avg_onsite_tat_hrs,
                "avg_rating": r.avg_rating,
                "avg_jobs_initiated": r.avg_jobs_initiated,
                "staffing": r.staffing,
                "performance": r.performance,
                "needs_coaching_count": r.needs_coaching_count,
                "strong_count": r.strong_count,
                "locations": r.locations or [],
                "solver_names": r.solver_names or [],
            }
            for r in rows
        ],
    }


# ---------------------------------------------------------------------------
# Period comparison — pick A and B, get per-solver and team deltas.
# Used both for the standalone comparison view and to enrich coaching docs.
# ---------------------------------------------------------------------------

@router.get("/api/periods/{a_id}/compare/{b_id}")
def compare_periods(
    a_id: int,
    b_id: int,
    db: Session = Depends(get_db),
    _: str = Depends(auth.require_admin),
):
    """Compare two periods. `a` is the more recent / current, `b` is the
    earlier / baseline."""
    a = db.get(models.Period, a_id)
    b = db.get(models.Period, b_id)
    if not a or not b:
        raise HTTPException(status_code=404, detail="One or both periods not found")

    a_snaps = {s.name: snapshot_to_dict(s) for s in a.snapshots}
    b_snaps = {s.name: snapshot_to_dict(s) for s in b.snapshots}

    # Team-level comparison
    team_a = {
        "median_response_tat_hrs": a.median_response_tat_hrs,
        "avg_onsite_tat_hrs": a.avg_onsite_tat_hrs,
        "submission_rate": a.team_submission_rate,
        "volume": a.total_valuations,
        "avg_rating": a.avg_rating,
        "assigned_count": a.total_jobs_assigned or 0,
        "jobs_initiated": a.total_jobs_initiated_by_solvers or 0,
    }
    team_b = {
        "median_response_tat_hrs": b.median_response_tat_hrs,
        "avg_onsite_tat_hrs": b.avg_onsite_tat_hrs,
        "submission_rate": b.team_submission_rate,
        "volume": b.total_valuations,
        "avg_rating": b.avg_rating,
        "assigned_count": b.total_jobs_assigned or 0,
        "jobs_initiated": b.total_jobs_initiated_by_solvers or 0,
    }
    team_cmp = analysis.compare_snapshots(team_a, team_b)

    # Per-solver — only solvers present in BOTH periods
    common = sorted(set(a_snaps.keys()) & set(b_snaps.keys()))
    per_solver = []
    for name in common:
        cmp = analysis.compare_snapshots(a_snaps[name], b_snaps[name])
        if cmp:
            per_solver.append({
                "name": name,
                "current_focus": a_snaps[name].get("focus_areas") or [],
                **cmp,
            })

    return {
        "a": {"id": a.id, "label": a.label},
        "b": {"id": b.id, "label": b.label},
        "team": team_cmp,
        "solvers": per_solver,
        "a_only": sorted(set(a_snaps.keys()) - set(b_snaps.keys())),
        "b_only": sorted(set(b_snaps.keys()) - set(a_snaps.keys())),
    }
